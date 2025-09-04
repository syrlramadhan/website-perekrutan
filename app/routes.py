from flask import render_template, url_for, jsonify, request, redirect, flash, send_file
from flask_login import login_required, login_user, logout_user, current_user

from flask_wtf import FlaskForm
from wtforms import StringField, SubmitField


from app import app, db
from app.models import calgot, Admin
from app.forms import CalgotForm

import os
from werkzeug.utils import secure_filename
import io
from datetime import datetime
import re
from markupsafe import escape


class LoginForm(FlaskForm):
	username = StringField('Username')
	password = StringField('Password')
	submit = SubmitField('Login')


@app.route('/')
@app.route('/index')
def index():
	return render_template('index.html')


@app.route('/login', methods=['GET','POST'])
def login():
	form = LoginForm()
	if current_user.is_authenticated:
		return redirect(url_for('pendaftar'))
	if form.validate_on_submit():
		user = Admin.query.filter_by(username=form.username.data).first()
		if user is None or not user.check_password(form.password.data):
			flash("Username atau Password Salah!!")
			return redirect(url_for('login'))
		flash("Berhasil Login")
		login_user(user)
		return redirect(url_for('pendaftar'))
	return render_template('login.html', form=form)
	
@app.route('/logout')
@login_required
def logout():
	logout_user()
	return redirect(url_for('login'))

@app.route('/newDaftar', methods=['GET','POST'])
def newDaftar():
	form = CalgotForm()
	return render_template('daftar2.html', form=form)


@app.route('/pendaftar')
@login_required
def pendaftar():
	return render_template('pendaftar.html')

@app.route('/daftar')
def daftar():
	return render_template('daftar2.html')



@app.route('/news')
def news():
	return render_template('news.html')


@app.route('/about')
def about():
	"""Render the About page."""
	return render_template('about.html')



@app.route('/api/getData', methods=['GET'])
def getData():
	try:
		print("DEBUG: getData route called")
		daftar_calgot = calgot.query.all()
		print(f"DEBUG: Found {len(daftar_calgot)} records")
		
		data_jsom = []
		for baris in daftar_calgot:
			print(f"DEBUG: Processing record {baris.id}: {baris.nama_lengkap}")
			# Handle foto path safely
			foto_path = baris.foto or ''
			# Only strip known prefixes; don't blindly chop the first N chars
			if foto_path.startswith('app/static/'):
				foto_path = foto_path[len('app/static/'):]
			elif foto_path.startswith('static/'):
				foto_path = foto_path[len('static/'):]
			# If no photo uploaded, return empty string so frontend uses onerror fallback
			if not foto_path or foto_path == 'tidak_ada.jpg':
				foto_path = ''
			
			data_jsom.append({
				'id': baris.id,
				'nama_lengkap': baris.nama_lengkap,
				'nama_panggilan': baris.nama_panggilan,
				'jenis_kelamin': baris.jenis_kelamin,
				'nomor_wa': baris.nomor_wa,
				'email': baris.email,
				'username_telegram': baris.username_telegram or '',
				'alamat': baris.alamat,
				'pilihan_tinggal': baris.pilihan_tinggal or '',
				'kampus': baris.kampus,
				'jurusan': baris.jurusan,
				'foto': foto_path,
				'tanggal': baris.timestamp.strftime('%d-%m-%Y %H:%M') if baris.timestamp else '',
				'alasan': baris.alasan
			})
		
		print(f"DEBUG: Returning {len(data_jsom)} records")
		return jsonify(data_jsom)
	except Exception as e:
		print(f"ERROR in getData: {e}")
		import traceback
		traceback.print_exc()
		return jsonify({'error': str(e)}), 500


@app.route('/api/postData', methods=['GET','POST'])
def postData():
	try:
		print("DEBUG: postData route called")
		print("DEBUG: Form data:", request.form)
		print("DEBUG: Files:", request.files)
		
		# Get form data
		nama_lengkap = (request.form.get('nama_lengkap') or '').strip()
		nama_panggilan = (request.form.get('nama_panggilan') or '').strip()
		jenis_kelamin = (request.form.get('jenis_kelamin') or '').strip()
		email = (request.form.get('email') or '').strip()
		nomor_wa = (request.form.get('nomor_wa') or '').strip()
		username_telegram = (request.form.get('username_telegram') or '').strip()
		alamat = (request.form.get('alamat') or '').strip()
		pilihan_tinggal = (request.form.get('pilihan_tinggal') or '').strip()
		kampus = (request.form.get('kampus') or '').strip()
		jurusan = (request.form.get('jurusan') or '').strip()
		alasan = (request.form.get('alasan') or '').strip()

		print(f"DEBUG: Received data - nama: {nama_lengkap}, email: {email}")
		print(f"DEBUG: New fields - telegram: {username_telegram}, pilihan_tinggal: {pilihan_tinggal}")

		# --- Server-side validation and sanitization ---
		# Required checks
		required = {
			'nama_lengkap': nama_lengkap,
			'nama_panggilan': nama_panggilan,
			'jenis_kelamin': jenis_kelamin,
			'email': email,
			'nomor_wa': nomor_wa,
			'alamat': alamat,
			'pilihan_tinggal': pilihan_tinggal,
			'kampus': kampus,
			'jurusan': jurusan,
			'alasan': alasan
		}
		for k, v in required.items():
			if not v:
				return jsonify({'status': 'error', 'message': f'Field {k} is required.'}), 400

		# Basic email validation
		if not re.match(r"^\S+@\S+\.\S+$", email):
			return jsonify({'status': 'error', 'message': 'Email format tidak valid.'}), 400

		# Phone validation (digits, plus, dash, space)
		if not re.match(r'^[0-9+\-\s]{6,20}$', nomor_wa):
			return jsonify({'status': 'error', 'message': 'Nomor WA tidak valid. Gunakan angka dan +/-.'}), 400

		# jenis_kelamin allowed values (normalize)
		jk_lower = jenis_kelamin.lower()
		if jk_lower in ('l', 'laki-laki', 'male', 'pria'):
			jenis_kelamin = 'Pria'
		elif jk_lower in ('p', 'perempuan', 'female', 'wanita'):
			jenis_kelamin = 'Wanita'
		else:
			return jsonify({'status': 'error', 'message': 'Pilihan jenis kelamin tidak valid.'}), 400

		# Length limits
		if len(nama_lengkap) > 200 or len(nama_panggilan) > 100:
			return jsonify({'status': 'error', 'message': 'Panjang nama terlalu panjang.'}), 400
		if len(alamat) > 1000 or len(alasan) > 3000:
			return jsonify({'status': 'error', 'message': 'Teks terlalu panjang.'}), 400
		if len(kampus) > 300 or len(jurusan) > 300:
			return jsonify({'status': 'error', 'message': 'Panjang kampus/jurusan terlalu panjang.'}), 400

		# Sanitize inputs roughly: escape HTML tags to avoid stored XSS
		nama_lengkap = escape(nama_lengkap)
		nama_panggilan = escape(nama_panggilan)
		username_telegram = escape(username_telegram)
		alamat = escape(alamat)
		pilihan_tinggal = escape(pilihan_tinggal)
		kampus = escape(kampus)
		jurusan = escape(jurusan)
		alasan = escape(alasan)

		# Handle photo (validate and save) - default
		photo_path = "tidak_ada.jpg"  # default (relative to static)
		if 'photo' in request.files:
			photo = request.files['photo']
			if photo and photo.filename:
				# Basic mimetype check
				mimetype = photo.mimetype or ''
				if not mimetype.startswith('image/'):
					return jsonify({'status': 'error', 'message': 'File harus berupa gambar.'}), 400
				# size check (read bytes then rewind)
				photo_stream = photo.stream
				photo_stream.seek(0, os.SEEK_END)
				size = photo_stream.tell()
				photo_stream.seek(0)
				MAX_BYTES = 5 * 1024 * 1024
				if size > MAX_BYTES:
					return jsonify({'status': 'error', 'message': 'File gambar terlalu besar (max 5MB).'}), 400
				# Ensure destination directory exists inside the app static folder
				dest_dir = os.path.join(app.root_path, 'static', 'foto_calgot')
				os.makedirs(dest_dir, exist_ok=True)
				# Secure filename and preserve extension
				orig_filename = secure_filename(photo.filename)
				base, ext = os.path.splitext(orig_filename)
				ext = ext.lower() if ext else '.jpg'
				if ext not in ('.jpg', '.jpeg', '.png'):
					ext = '.jpg'
				# Use cleaned nama_lengkap as filename if available
				name_part = secure_filename(nama_lengkap) if nama_lengkap else base or 'photo'
				save_name = f"{name_part}{ext}"
				full_path = os.path.join(dest_dir, save_name)
				# Try to save file; if saving fails due to missing dir or permission, attempt to create dir and retry
				try:
					photo.save(full_path)
				except Exception as save_e:
					print(f"WARN: initial photo.save failed: {save_e}; attempting to create dir and retry")
					try:
						os.makedirs(dest_dir, exist_ok=True)
						photo.stream.seek(0)
						photo.save(full_path)
					except Exception as retry_e:
						print(f"ERROR: failed to save uploaded photo after retry: {retry_e}")
						return jsonify({'status': 'error', 'message': 'Gagal menyimpan file foto. Pastikan direktori server dapat ditulisi.'}), 500
				# Store path relative to static folder (so frontend can use /static/<foto>)
				photo_path = os.path.join('foto_calgot', save_name)
				print(f"DEBUG: Photo saved to {full_path}")
		
		# Create calgot object
		add_calgot = calgot(
			nama_lengkap=nama_lengkap,
			nama_panggilan=nama_panggilan,
			jenis_kelamin=jenis_kelamin,
			email=email,
			nomor_wa=nomor_wa,
			username_telegram=username_telegram,
			alamat=alamat,
			pilihan_tinggal=pilihan_tinggal,
			kampus=kampus,
			jurusan=jurusan,
			foto=photo_path,
			alasan=alasan
		)
		
		print(f"DEBUG: Creating calgot object: {add_calgot}")
		
		# Save to database
		db.session.add(add_calgot)
		db.session.commit()
		
		print(f"DEBUG: Calgot {nama_lengkap} berhasil didaftarkan")
		return jsonify({'status': 'success', 'message': 'Pendaftaran berhasil!'})
		
	except Exception as e:
		print(f"ERROR in postData: {e}")
		import traceback
		traceback.print_exc()
		db.session.rollback()
		return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/pendaftar/export/xlsx')
@login_required
def export_pendaftar_xlsx():
	"""Export daftar pendaftar sebagai file XLSX with improved styling and column sizing."""
	try:
		daftar_calgot = calgot.query.order_by(calgot.id).all()
		try:
			from openpyxl import Workbook
			from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
			from openpyxl.utils import get_column_letter
		except Exception as imp_e:
			return jsonify({'error': 'openpyxl library is required to export XLSX. Install with: pip install openpyxl', 'details': str(imp_e)}), 500

		wb = Workbook()
		ws = wb.active
		ws.title = 'Pendaftar'

		headers = [
			'ID', 'Nama Lengkap', 'Nama Panggilan', 'Jenis Kelamin', 'Email', 'Nomor WA',
			'Username Telegram', 'Alamat', 'Pilihan Tinggal', 'Kampus', 'Jurusan', 'Alasan', 'Tanggal'
		]
		ws.append(headers)

		# Styles
		header_font = Font(bold=True, color='FFFFFF')
		header_fill = PatternFill('solid', fgColor='244a9b')
		thin = Side(border_style='thin', color='000000')
		border = Border(left=thin, right=thin, top=thin, bottom=thin)

		for col_num, header in enumerate(headers, 1):
			cell = ws.cell(row=1, column=col_num)
			cell.font = header_font
			cell.fill = header_fill
			cell.alignment = Alignment(horizontal='center', wrap_text=True)
			cell.border = border

		# Append rows
		for row in daftar_calgot:
			foto_path = row.foto or ''
			if foto_path.startswith('app/static/'):
				foto_path = foto_path[len('app/static/'):]
			ws.append([
				row.id,
				row.nama_lengkap or '',
				row.nama_panggilan or '',
				row.jenis_kelamin or '',
				row.email or '',
				row.nomor_wa or '',
				row.username_telegram or '',
				row.alamat or '',
				row.pilihan_tinggal or '',
				row.kampus or '',
				row.jurusan or '',
				row.alasan or '',
				row.timestamp.strftime('%Y-%m-%d %H:%M:%S') if row.timestamp else ''
			])

		# Compute column widths based on max length per column
		dims = {}
		for row in ws.rows:
			for cell in row:
				if cell.value is None:
					length = 0
				else:
					length = len(str(cell.value))
				col = cell.column_letter
				dims[col] = max(dims.get(col, 0), length)

		for col, length in dims.items():
			# heuristic: add some padding, clamp to reasonable range
			width = min(max(length + 4, 15), 60)
			ws.column_dimensions[col].width = width

		# Apply alignment/wrap and borders to data cells
		for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
			for cell in row:
				if headers[cell.column - 1] in ('Alamat', 'Alasan'):
					cell.alignment = Alignment(wrap_text=True, vertical='top')
				else:
					cell.alignment = Alignment(vertical='top')
				cell.border = border

		# Freeze header row and enable auto filter
		ws.freeze_panes = 'A2'
		ws.auto_filter.ref = ws.dimensions

		output = io.BytesIO()
		wb.save(output)
		output.seek(0)

		filename = f'pendaftar_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.xlsx'
		return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	except Exception as e:
		import traceback
		traceback.print_exc()
		return jsonify({'error': str(e)}), 500


@app.route('/pendaftar/export/pdf')
@login_required
def export_pendaftar_pdf():
	"""Export daftar pendaftar sebagai PDF with landscape layout and wrapped cells."""
	try:
		daftar_calgot = calgot.query.order_by(calgot.id).all()
		try:
			from reportlab.lib.pagesizes import A4, landscape
			from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
			from reportlab.lib import colors
			from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
		except Exception as imp_e:
			return jsonify({'error': 'reportlab library is required to export PDF. Install with: pip install reportlab', 'details': str(imp_e)}), 500

		output = io.BytesIO()
		page_size = landscape(A4)
		doc = SimpleDocTemplate(output, pagesize=page_size, rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
		elements = []
		styles = getSampleStyleSheet()
		title_style = ParagraphStyle('title', parent=styles['Title'], alignment=1, fontSize=16)
		small = ParagraphStyle('small', parent=styles['BodyText'], fontSize=8, leading=10)

		title = Paragraph('Daftar Pendaftar - COCONUT Computer Club', title_style)
		elements.append(title)
		elements.append(Spacer(1, 8))

		headers = ['ID', 'Nama Lengkap', 'Panggilan', 'JK', 'Email', 'WA', 'Telegram', 'Kampus', 'Jurusan', 'Pilihan Tinggal', 'Alasan', 'Tanggal']

		data = [headers]
		for r in daftar_calgot:
			data.append([
				str(r.id),
				Paragraph(r.nama_lengkap or '-', small),
				Paragraph(r.nama_panggilan or '-', small),
				r.jenis_kelamin or '-',
				Paragraph(r.email or '-', small),
				r.nomor_wa or '-',
				r.username_telegram or '-',
				Paragraph((r.kampus or '-')[:80], small),
				Paragraph((r.jurusan or '-')[:60], small),
				r.pilihan_tinggal or '-',
				Paragraph((r.alasan or '-')[:1000], small),
				r.timestamp.strftime('%Y-%m-%d') if r.timestamp else ''
			])

		# Estimate column widths to fit landscape A4 (usable width ~ page_size[0] - margins)
		usable_width = page_size[0] - 40  # left+right margins total 40
		col_widths = [40, 140, 80, 30, 140, 70, 70, 120, 80, 70, 220, 60]
		# normalize if sums differ
		total = sum(col_widths)
		if total != usable_width:
			scale = usable_width / total
			col_widths = [w * scale for w in col_widths]

		table = Table(data, colWidths=col_widths, repeatRows=1)
		table_style = TableStyle([
			('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#244a9b')),
			('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
			('ALIGN', (0, 0), (-1, 0), 'CENTER'),
			('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
			('FONTSIZE', (0, 0), (-1, -1), 8),
			('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
			('VALIGN', (0, 0), (-1, -1), 'TOP'),
			('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey])
		])
		table.setStyle(table_style)

		elements.append(table)
		doc.build(elements)
		output.seek(0)

		filename = f'pendaftar_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.pdf'
		return send_file(output, as_attachment=True, download_name=filename, mimetype='application/pdf')
	except Exception as e:
		import traceback
		traceback.print_exc()
		return jsonify({'error': str(e)}), 500
